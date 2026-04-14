"""XLSX 多 sheet 原始数据提取 (C5 parser-pipeline)

openpyxl 读所有 sheet(含隐藏),返回:
- 每 sheet 一条合并文本(喂相似度)
- 每 sheet 原始 cell 矩阵(供后续 LLM 报价规则识别 + fill_price 回填)
"""

from __future__ import annotations

import logging
from dataclasses import dataclass, field
from pathlib import Path

logger = logging.getLogger(__name__)


@dataclass(frozen=True)
class SheetData:
    sheet_name: str
    hidden: bool
    # 合并文本(非空 cell 按行拼接,|分隔)
    merged_text: str
    # 原始矩阵:rows[row_idx][col_idx] = cell value(str | float | None)
    rows: list[list[object]]
    # 合并单元格 ranges 字符串列表(C9 structure_similarity 字段维度消费)
    # 如 ["A1:B2", "C3:D4"],openpyxl ws.merged_cells.ranges 的 str(r) 结果
    merged_cells_ranges: list[str] = field(default_factory=list)


@dataclass(frozen=True)
class XlsxExtractResult:
    sheets: list[SheetData] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)


def extract_xlsx(file_path: str | Path) -> XlsxExtractResult:
    """从 xlsx 抽所有 sheet。异常向上抛。"""
    from openpyxl import load_workbook

    wb = load_workbook(str(file_path), data_only=True, read_only=False)
    sheets: list[SheetData] = []
    warnings: list[str] = []

    # Workbook.sheetnames 默认不含隐藏,用 _sheets 全量遍历
    for ws in wb.worksheets:
        hidden = ws.sheet_state != "visible"
        try:
            rows: list[list[object]] = []
            cells_text: list[str] = []
            for row in ws.iter_rows(values_only=True):
                row_list: list[object] = [cell for cell in row]
                rows.append(row_list)
                row_text = " | ".join(
                    str(c).strip()
                    for c in row_list
                    if c is not None and str(c).strip()
                )
                if row_text:
                    cells_text.append(row_text)
            merged = "\n".join(cells_text)
            # 合并单元格 ranges(C9):ws.merged_cells.ranges 是 MergedCellRange 对象,
            # str(r) → "A1:B2" 形式字符串
            try:
                merged_ranges = [str(r) for r in ws.merged_cells.ranges]
            except Exception:  # pragma: no cover - openpyxl 极端兼容
                merged_ranges = []
            sheets.append(
                SheetData(
                    sheet_name=ws.title,
                    hidden=hidden,
                    merged_text=merged,
                    rows=rows,
                    merged_cells_ranges=merged_ranges,
                )
            )
        except Exception as e:  # pragma: no cover - 单 sheet 失败隔离
            msg = f"sheet {ws.title!r} extract failed: {e!s}"[:200]
            warnings.append(msg)
            logger.warning(msg)

    wb.close()
    return XlsxExtractResult(sheets=sheets, warnings=warnings)


__all__ = ["extract_xlsx", "XlsxExtractResult", "SheetData"]
