"""L2 - C5 run_pipeline 全链路 + fill_price 回填 + 状态机终态 (spec Req 4+5+8)

覆盖:
- pipeline 完整路径: extracted → identifying → identified → pricing → priced
- 无报价表 bidder 停在 identified
- 内容提取全失败 → identify_failed
- 报价规则识别失败 → price_failed
- fill_price sheet 回填、千分位、空行、terminal 判定 (priced / price_partial / price_failed)
"""

from __future__ import annotations

import os
from pathlib import Path

import pytest_asyncio
from sqlalchemy import select

from app.db.session import async_session
from app.models.bid_document import BidDocument
from app.models.bidder import Bidder
from app.models.price_item import PriceItem
from app.models.price_parsing_rule import PriceParsingRule
from app.models.project import Project
from app.models.user import User
from app.services.auth.password import hash_password
from app.services.parser.pipeline import rule_coordinator
from app.services.parser.pipeline.run_pipeline import run_pipeline
from tests.fixtures.auth_fixtures import clean_users as _clean_users  # noqa
from tests.fixtures.doc_fixtures import make_price_xlsx, make_real_docx
from tests.fixtures.llm_mock import (
    ScriptedLLMProvider,
    make_price_rule_response,
    make_role_classify_response,
)

os.environ.setdefault("INFRA_DISABLE_PIPELINE", "1")


async def _seed_project_with_bidder(docs: list[dict]) -> tuple[int, int, list[int]]:
    """docs: [{"name": "x.docx", "type": ".docx", "path": Path}, ...]
    返回 (project_id, bidder_id, doc_ids)
    """
    async with async_session() as s:
        user = User(
            username="px",
            password_hash=hash_password("x"),
            role="reviewer",
            must_change_password=False,
        )
        s.add(user)
        await s.flush()
        project = Project(name="P", owner_id=user.id)
        s.add(project)
        await s.flush()
        bidder = Bidder(
            name="B", project_id=project.id, parse_status="extracted"
        )
        s.add(bidder)
        await s.flush()
        ids: list[int] = []
        for i, spec in enumerate(docs):
            doc = BidDocument(
                bidder_id=bidder.id,
                file_name=spec["name"],
                file_path=str(spec["path"]),
                file_size=1000,
                file_type=spec["type"],
                md5=(f"{i:02d}" + "p" * 30)[:32],
                source_archive="a.zip",
                parse_status="extracted",
            )
            s.add(doc)
            await s.flush()
            ids.append(doc.id)
        await s.commit()
        return project.id, bidder.id, ids


async def test_pipeline_happy_path_priced(clean_users, tmp_path: Path):
    rule_coordinator.reset_for_tests()
    docx_path = make_real_docx(tmp_path / "t.docx", body_paragraphs=["技术内容"])
    xlsx_path = make_price_xlsx(tmp_path / "p.xlsx", row_count=3)
    pid, bid, ids = await _seed_project_with_bidder(
        [
            {"name": "技术方案.docx", "type": ".docx", "path": docx_path},
            {"name": "投标报价.xlsx", "type": ".xlsx", "path": xlsx_path},
        ]
    )
    llm = ScriptedLLMProvider(
        [
            make_role_classify_response(
                [(ids[0], "technical"), (ids[1], "pricing")],
                identity_info={"company_full_name": "A 公司"},
            ),
            make_price_rule_response(),
        ]
    )

    await run_pipeline(bid, llm=llm)

    async with async_session() as s:
        bidder = await s.get(Bidder, bid)
        assert bidder.parse_status == "priced"
        items = (
            await s.execute(select(PriceItem).where(PriceItem.bidder_id == bid))
        ).scalars().all()
        assert len(items) == 3


async def test_pipeline_no_pricing_stops_at_identified(
    clean_users, tmp_path: Path
):
    rule_coordinator.reset_for_tests()
    docx = make_real_docx(tmp_path / "t.docx", body_paragraphs=["x"])
    pid, bid, ids = await _seed_project_with_bidder(
        [{"name": "技术.docx", "type": ".docx", "path": docx}]
    )
    llm = ScriptedLLMProvider(
        [make_role_classify_response([(ids[0], "technical")])]
    )
    await run_pipeline(bid, llm=llm)

    async with async_session() as s:
        bidder = await s.get(Bidder, bid)
        assert bidder.parse_status == "identified"
        # 无 price_parsing_rules
        rules = (
            await s.execute(select(PriceParsingRule).where(PriceParsingRule.project_id == pid))
        ).scalars().all()
        assert rules == []


async def test_pipeline_all_content_failed_identify_failed(
    clean_users, tmp_path: Path
):
    rule_coordinator.reset_for_tests()
    bad = tmp_path / "bad.docx"
    bad.write_bytes(b"not valid docx")
    pid, bid, _ = await _seed_project_with_bidder(
        [{"name": "bad.docx", "type": ".docx", "path": bad}]
    )
    llm = ScriptedLLMProvider([make_role_classify_response([])])
    await run_pipeline(bid, llm=llm)
    async with async_session() as s:
        bidder = await s.get(Bidder, bid)
        assert bidder.parse_status == "identify_failed"


async def test_pipeline_price_rule_failure_leads_to_price_failed(
    clean_users, tmp_path: Path
):
    from app.services.llm.base import LLMError

    rule_coordinator.reset_for_tests()
    docx = make_real_docx(tmp_path / "t.docx", body_paragraphs=["x"])
    xlsx = make_price_xlsx(tmp_path / "p.xlsx")
    pid, bid, ids = await _seed_project_with_bidder(
        [
            {"name": "技术.docx", "type": ".docx", "path": docx},
            {"name": "投标报价.xlsx", "type": ".xlsx", "path": xlsx},
        ]
    )
    # 第一次 LLM 调用(角色)成功;第二次(规则识别)失败
    llm = ScriptedLLMProvider(
        [
            make_role_classify_response(
                [(ids[0], "technical"), (ids[1], "pricing")]
            ),
            LLMError(kind="timeout", message="rule fail"),
        ],
        loop_last=True,
    )
    await run_pipeline(bid, llm=llm)
    async with async_session() as s:
        bidder = await s.get(Bidder, bid)
        assert bidder.parse_status == "price_failed"


async def test_pipeline_thousand_sep_and_currency_normalized(
    clean_users, tmp_path: Path
):
    """构造含千分位值的 xlsx,回填后 unit_price 为标准 Decimal。"""
    import openpyxl

    rule_coordinator.reset_for_tests()
    docx = make_real_docx(tmp_path / "t.docx", body_paragraphs=["x"])
    xlsx = tmp_path / "money.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("报价清单")
    ws.append(["标题", "", "", "", "", ""])
    ws.append(["编码", "名称", "单位", "数量", "单价", "合价"])
    ws.append(["A1", "项1", "m3", "10", "1,234.56", "12,345.60"])
    wb.save(str(xlsx))

    pid, bid, ids = await _seed_project_with_bidder(
        [
            {"name": "技术.docx", "type": ".docx", "path": docx},
            {"name": "投标报价.xlsx", "type": ".xlsx", "path": xlsx},
        ]
    )
    llm = ScriptedLLMProvider(
        [
            make_role_classify_response(
                [(ids[0], "technical"), (ids[1], "pricing")]
            ),
            make_price_rule_response(),
        ]
    )
    await run_pipeline(bid, llm=llm)

    async with async_session() as s:
        items = (
            await s.execute(select(PriceItem).where(PriceItem.bidder_id == bid))
        ).scalars().all()
        assert len(items) == 1
        item = items[0]
        assert str(item.unit_price) == "1234.56"
        assert str(item.total_price) == "12345.60"


# ============================================================================
# fix-unit-price-orphan-fallback: L2 e2e for unit_price fallback + invariant
# ============================================================================


async def _seed_project_with_n_bidders(
    n: int, docs_per_bidder: list[list[dict]],
    project_max_price: float | None = None,
) -> tuple[int, list[int], list[list[int]]]:
    """seed 1 project + N bidders + per-bidder docs。

    docs_per_bidder[i] = bidder i 的 doc spec 列表(同 _seed_project_with_bidder)
    返回 (project_id, [bidder_ids], [[doc_ids per bidder], ...])
    """
    from decimal import Decimal

    assert len(docs_per_bidder) == n
    async with async_session() as s:
        user = User(
            username=f"upf_{n}_{id(docs_per_bidder)}",
            password_hash=hash_password("x"),
            role="reviewer",
            must_change_password=False,
        )
        s.add(user)
        await s.flush()
        project = Project(
            name=f"upf_P_{n}",
            owner_id=user.id,
            max_price=Decimal(str(project_max_price)) if project_max_price else None,
        )
        s.add(project)
        await s.flush()

        bidder_ids: list[int] = []
        all_doc_ids: list[list[int]] = []
        for i in range(n):
            bidder = Bidder(
                name=f"upf_B{i}",
                project_id=project.id,
                parse_status="extracted",
            )
            s.add(bidder)
            await s.flush()
            bidder_ids.append(bidder.id)
            doc_ids: list[int] = []
            for j, spec in enumerate(docs_per_bidder[i]):
                doc = BidDocument(
                    bidder_id=bidder.id,
                    file_name=spec["name"],
                    file_path=str(spec["path"]),
                    file_size=1000,
                    file_type=spec["type"],
                    md5=(f"upf{i:02d}{j:02d}" + "x" * 24)[:32],
                    source_archive=f"upf-b{i}.zip",
                    parse_status="extracted",
                )
                s.add(doc)
                await s.flush()
                doc_ids.append(doc.id)
            all_doc_ids.append(doc_ids)
        await s.commit()
        return project.id, bidder_ids, all_doc_ids


async def test_pipeline_unit_price_fallback(clean_users, tmp_path: Path):
    """L2:2 个 bidder 的 xlsx 全被 LLM 误判为 unit_price → fallback 兜底回填。

    覆盖 spec ADDED Scenario "仅 unit_price 类 XLSX(fallback)" + "项目内
    bidder 落到不同 role(每家独立判定)"。验证:
    - 两家 bidder 都进 priced 终态(不卡 identified)
    - 各自 price_items 行数 ≥ 1
    - aggregate_bidder_totals 返 2 条,无重复求和

    设计取 2 家而非任务原文的 3 家:多 bidder 聚合不变量在 2 即可暴露;
    3 家场景在 L3 真实 UI walkthrough 覆盖(更接近生产症状)。
    """
    from app.services.detect.agents.anomaly_impl.config import AnomalyConfig
    from app.services.detect.agents.anomaly_impl.extractor import (
        aggregate_bidder_totals,
    )

    rule_coordinator.reset_for_tests()
    docx_a = make_real_docx(tmp_path / "ta.docx", body_paragraphs=["技术 A"])
    xlsx_a = make_price_xlsx(tmp_path / "pa.xlsx", row_count=2)
    docx_b = make_real_docx(tmp_path / "tb.docx", body_paragraphs=["技术 B"])
    xlsx_b = make_price_xlsx(tmp_path / "pb.xlsx", row_count=2)

    pid, bids, doc_ids = await _seed_project_with_n_bidders(
        n=2,
        docs_per_bidder=[
            [
                {"name": "技术.docx", "type": ".docx", "path": docx_a},
                {"name": "工程监理报价表.xlsx", "type": ".xlsx", "path": xlsx_a},
            ],
            [
                {"name": "技术.docx", "type": ".docx", "path": docx_b},
                {"name": "附件5 工程监理报价表.xlsx", "type": ".xlsx", "path": xlsx_b},
            ],
        ],
    )

    # bidder A: leader → role_classify + price_rule;LLM 判 xlsx 为 unit_price
    llm_a = ScriptedLLMProvider(
        [
            make_role_classify_response(
                [(doc_ids[0][0], "technical"), (doc_ids[0][1], "unit_price")]
            ),
            make_price_rule_response(),
        ]
    )
    await run_pipeline(bids[0], llm=llm_a)

    # bidder B: 用同项目 confirmed rule;只需 role_classify
    llm_b = ScriptedLLMProvider(
        [
            make_role_classify_response(
                [(doc_ids[1][0], "technical"), (doc_ids[1][1], "unit_price")]
            ),
        ]
    )
    await run_pipeline(bids[1], llm=llm_b)

    # 关键断言:两家 bidder 全 priced(fallback 生效),各 price_items 非空
    async with async_session() as s:
        b_a = await s.get(Bidder, bids[0])
        b_b = await s.get(Bidder, bids[1])
        assert b_a.parse_status == "priced", f"bidder A 应 priced, 实际={b_a.parse_status}"
        assert b_b.parse_status == "priced", f"bidder B 应 priced, 实际={b_b.parse_status}"

        items_a = (await s.execute(
            select(PriceItem).where(PriceItem.bidder_id == bids[0])
        )).scalars().all()
        items_b = (await s.execute(
            select(PriceItem).where(PriceItem.bidder_id == bids[1])
        )).scalars().all()
        assert len(items_a) >= 1, "bidder A 应有 price_items"
        assert len(items_b) >= 1, "bidder B 应有 price_items"

        # aggregate 不重复求和:返 2 条记录
        cfg = AnomalyConfig()
        summaries = await aggregate_bidder_totals(s, pid, cfg)
        assert len(summaries) == 2, (
            f"aggregate 应返 2 条 BidderPriceSummary, 实际={len(summaries)}"
        )


async def test_pipeline_mixed_role_no_double_count(clean_users, tmp_path: Path):
    """L2:1 个 bidder 同时上传 pricing + unit_price 两份 xlsx → 仅 pricing 进回填。

    覆盖 spec ADDED Scenario "pricing + unit_price 都有(优先 pricing 不混合)"。
    集成验证:跑 price_overshoot detector,断言 score=0(不变量真的拦住了
    "主表+子表混算→误报铁证升 high")。

    设值:pricing.xlsx 总价 ≈ 100;unit_price.xlsx 总价 ≈ 50;max_price=120
    - 若混算 SUM=150 > 120 → price_overshoot 误报铁证
    - 若不变量生效 SUM=100 < 120 → score=0 ✓
    """
    import openpyxl
    from decimal import Decimal

    from app.models.agent_task import AgentTask
    from app.services.detect.agents.price_overshoot import (
        run as price_overshoot_run,
    )
    from app.services.detect.agents.anomaly_impl.config import AnomalyConfig
    from app.services.detect.agents.anomaly_impl.extractor import (
        aggregate_bidder_totals,
    )
    from app.services.detect.context import AgentContext

    rule_coordinator.reset_for_tests()

    # pricing.xlsx 行总价 = 100(用 make_price_xlsx 自带的 row_count=1 + total=100)
    pricing_xlsx = tmp_path / "pricing.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("报价清单")
    ws.append(["标题", "", "", "", "", ""])
    ws.append(["编码", "名称", "单位", "数量", "单价", "合价"])
    ws.append(["A1", "项 P", "m3", "1", "100", "100"])
    wb.save(str(pricing_xlsx))

    # unit_price.xlsx 同结构,行总价 = 50
    up_xlsx = tmp_path / "unit_price.xlsx"
    wb2 = openpyxl.Workbook()
    wb2.remove(wb2.active)
    ws2 = wb2.create_sheet("报价清单")
    ws2.append(["标题", "", "", "", "", ""])
    ws2.append(["编码", "名称", "单位", "数量", "单价", "合价"])
    ws2.append(["U1", "项 U", "m3", "1", "50", "50"])
    wb2.save(str(up_xlsx))

    docx = make_real_docx(tmp_path / "t.docx", body_paragraphs=["技术"])

    pid, bids, doc_ids = await _seed_project_with_n_bidders(
        n=1,
        docs_per_bidder=[
            [
                {"name": "技术.docx", "type": ".docx", "path": docx},
                {"name": "主报价表.xlsx", "type": ".xlsx", "path": pricing_xlsx},
                {"name": "综合单价分析表.xlsx", "type": ".xlsx", "path": up_xlsx},
            ],
        ],
        project_max_price=120.0,  # 主表单算不超 / 混算超
    )

    # LLM mock:doc 1 → pricing, doc 2 → unit_price
    llm = ScriptedLLMProvider(
        [
            make_role_classify_response(
                [
                    (doc_ids[0][0], "technical"),
                    (doc_ids[0][1], "pricing"),
                    (doc_ids[0][2], "unit_price"),
                ]
            ),
            make_price_rule_response(),  # 只对 pricing xlsx 调一次
        ]
    )
    await run_pipeline(bids[0], llm=llm)

    async with async_session() as s:
        bidder = await s.get(Bidder, bids[0])
        assert bidder.parse_status == "priced"

        items = (await s.execute(
            select(PriceItem).where(PriceItem.bidder_id == bids[0])
        )).scalars().all()
        # 不变量:仅来自 pricing.xlsx 的 1 行(item_code='A1'),不含 unit_price 的 'U1'
        assert len(items) == 1, f"应只 1 行(来自 pricing),实际 {len(items)}"
        assert items[0].item_code == "A1", (
            f"应只含 pricing xlsx 的行(A1),实际 item_code={items[0].item_code}"
        )

        # aggregate 总价应 = 100(pricing 单算),不是 150(混算)
        cfg = AnomalyConfig()
        summaries = await aggregate_bidder_totals(s, pid, cfg)
        assert len(summaries) == 1
        assert summaries[0]["total_price"] == 100.0, (
            f"应 SUM=100(仅 pricing),实际={summaries[0]['total_price']}"
        )

        # 集成验证:跑 price_overshoot,score 应 = 0(120 > 100,未超限)
        agent_task = AgentTask(
            project_id=pid,
            version=1,
            agent_name="price_overshoot",
            agent_type="global",
            status="running",
        )
        s.add(agent_task)
        await s.flush()
        ctx = AgentContext(
            project_id=pid,
            version=1,
            agent_task=agent_task,
            bidder_a=None,
            bidder_b=None,
            all_bidders=[bidder],
            session=s,
        )
        result = await price_overshoot_run(ctx)
        await s.commit()

        assert result.score == 0.0, (
            f"price_overshoot 误报! score={result.score}, summary={result.summary}; "
            f"如果 = 100 说明混算未被拦住"
        )
        assert result.evidence_json["has_iron_evidence"] is False, (
            "不应有 iron evidence(120 > pricing-only SUM 100)"
        )
