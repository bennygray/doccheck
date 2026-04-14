"""L1 - xlsx_parser merged_cells_ranges 字段 (C9 §2)

覆盖:
- 有合并单元格 → merged_cells_ranges 含对应字符串
- 无合并单元格 → merged_cells_ranges 为空 list
- 多 sheet 合并单元格分别归属各 sheet(不混)
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from app.services.parser.content.xlsx_parser import extract_xlsx


def _make_xlsx_with_merged(
    out: Path,
    sheets: dict[str, tuple[list[list[object]], list[str]]],
) -> Path:
    """sheets[name] = (rows, merged_ranges_list)"""
    wb = Workbook()
    default = wb.active
    wb.remove(default)
    for name, (rows, merges) in sheets.items():
        ws = wb.create_sheet(title=name)
        for row in rows:
            ws.append(row)
        for rng in merges:
            ws.merge_cells(rng)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out))
    return out


def test_merged_cells_captured(tmp_path: Path) -> None:
    path = _make_xlsx_with_merged(
        tmp_path / "a.xlsx",
        sheets={
            "数据": (
                [["A", "B", "C"], [1, 2, 3], [4, 5, 6]],
                ["A1:C1", "A2:B2"],
            ),
        },
    )
    result = extract_xlsx(path)
    assert len(result.sheets) == 1
    s = result.sheets[0]
    # set 去顺序比较(openpyxl ranges 顺序不保证)
    assert set(s.merged_cells_ranges) == {"A1:C1", "A2:B2"}


def test_no_merged_cells_returns_empty(tmp_path: Path) -> None:
    path = _make_xlsx_with_merged(
        tmp_path / "b.xlsx",
        sheets={"Plain": ([["x", "y"], [1, 2]], [])},
    )
    result = extract_xlsx(path)
    s = result.sheets[0]
    assert s.merged_cells_ranges == []


def test_multi_sheet_merged_cells_isolated(tmp_path: Path) -> None:
    path = _make_xlsx_with_merged(
        tmp_path / "c.xlsx",
        sheets={
            "S1": ([["h1", "h2"], [1, 2]], ["A1:B1"]),
            "S2": ([["x", "y", "z"], [1, 2, 3]], ["A1:C1", "A2:B2"]),
        },
    )
    result = extract_xlsx(path)
    by_name = {s.sheet_name: s for s in result.sheets}
    assert set(by_name["S1"].merged_cells_ranges) == {"A1:B1"}
    assert set(by_name["S2"].merged_cells_ranges) == {"A1:C1", "A2:B2"}
