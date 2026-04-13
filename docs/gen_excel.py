"""生成第一期层级需求清单 Excel 文件"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ============================================================
# 数据定义
# ============================================================
MODULES = [
    ("AUTH", "认证与授权"),
    ("PROJ", "项目管理"),
    ("FILE", "投标人与文件管理"),
    ("PARSE", "文档解析"),
    ("DETECT", "检测执行"),
    ("REPORT", "报告与结果"),
    ("COMPARE", "对比视图"),
    ("ADMIN", "管理后台"),
    ("INFRA", "技术基础设施"),
]

# (编号, 任务, 归属, 工作量, 关联US, 关联API, 关联表, AC数, 模块, 功能项)
TASKS = [
    # === AUTH ===
    ("AUTH-01-BE1", "实现 POST /api/auth/login 接口（凭证校验、JWT 生成、失败计数、账户锁定）", "BE", "中", "US-1.1", "POST /api/auth/login", "User", 5, "AUTH", "AUTH-01 用户登录"),
    ("AUTH-01-BE2", "实现 GET /api/auth/me 接口（解析 JWT、返回用户信息含 must_change_password）", "BE", "小", "US-1.1", "GET /api/auth/me", "User", 2, "AUTH", "AUTH-01 用户登录"),
    ("AUTH-01-BE3", "JWT 中间件依赖注入（get_current_user），所有受保护路由集成", "BE", "中", "US-1.1", "全部受保护端点", "User", 2, "AUTH", "AUTH-01 用户登录"),
    ("AUTH-01-FE1", "登录页面（/login）：表单、错误提示、loading 状态", "FE", "小", "US-1.1", "POST /api/auth/login", "-", 3, "AUTH", "AUTH-01 用户登录"),
    ("AUTH-01-FE2", "Token 管理：localStorage 存取 + axios 拦截器自动附带 Authorization header", "FE", "小", "US-1.1", "-", "-", 2, "AUTH", "AUTH-01 用户登录"),
    ("AUTH-01-FE3", "401 拦截器：token 过期提示 + 表单防丢失（缓存未提交数据到 localStorage）", "FE", "中", "US-1.3", "-", "-", 2, "AUTH", "AUTH-01 用户登录"),
    ("AUTH-02-FE1", "登出按钮（顶部栏）：清除 localStorage token + 跳转 /login", "FE", "小", "US-1.2", "-", "-", 3, "AUTH", "AUTH-02 用户登出"),
    ("AUTH-03-FE1", "ProtectedRoute 组件：未登录→/login，审查员访问 /admin→/projects", "FE", "中", "US-1.3", "GET /api/auth/me", "-", 4, "AUTH", "AUTH-03 路由守卫"),
    ("AUTH-03-BE1", "后端每个 API 端点权限装饰器（角色校验：reviewer/admin）", "BE", "中", "US-1.3", "全部端点", "User", 3, "AUTH", "AUTH-03 路由守卫"),
    ("AUTH-03-BE2", "管理员对他人项目只读访问逻辑", "BE", "小", "US-1.3", "项目相关端点", "Project", 2, "AUTH", "AUTH-03 路由守卫"),
    ("AUTH-04-BE1", "数据库初始化脚本创建默认管理员（admin/admin123, must_change_password=true）", "BE", "小", "US-1.4", "-", "User", 2, "AUTH", "AUTH-04 初始用户与改密"),
    ("AUTH-04-BE2", "POST /api/auth/change-password 接口（旧密码校验、新密码规则）", "BE", "小", "US-1.4", "POST /api/auth/change-password", "User", 4, "AUTH", "AUTH-04 初始用户与改密"),
    ("AUTH-04-FE1", "修改密码页面 + ForceChangePassword 拦截组件", "FE", "小", "US-1.4", "POST /api/auth/change-password", "-", 3, "AUTH", "AUTH-04 初始用户与改密"),

    # === PROJ ===
    ("PROJ-01-BE1", "POST /api/projects 接口（字段校验、owner 关联、status=draft）", "BE", "小", "US-2.1", "POST /api/projects", "Project", 5, "PROJ", "PROJ-01 创建项目"),
    ("PROJ-01-FE1", "创建项目页面（/projects/new）：居中表单、字段校验、限价提示文案", "FE", "小", "US-2.1", "POST /api/projects", "-", 3, "PROJ", "PROJ-01 创建项目"),
    ("PROJ-02-BE1", "GET /api/projects 接口（分页、筛选、搜索、权限过滤）", "BE", "中", "US-2.2", "GET /api/projects", "Project", 5, "PROJ", "PROJ-02 项目列表"),
    ("PROJ-02-FE1", "项目列表页面（/projects）：卡片网格、筛选栏、搜索框、空状态引导", "FE", "中", "US-2.2", "GET /api/projects", "-", 5, "PROJ", "PROJ-02 项目列表"),
    ("PROJ-03-BE1", "GET /api/projects/{id} 接口（含 bidders[] + reports[] + current_analysis?）", "BE", "中", "US-2.3", "GET /api/projects/{id}", "Project,Bidder,AnalysisReport,AgentTask", 6, "PROJ", "PROJ-03 项目详情"),
    ("PROJ-03-FE1", "项目详情页面：项目信息栏 + 投标人管理区 + 检测记录区", "FE", "大", "US-2.3", "GET /api/projects/{id}", "-", 6, "PROJ", "PROJ-03 项目详情"),
    ("PROJ-03-FE2", "状态流转展示 + 按钮启禁逻辑", "FE", "小", "US-2.3", "-", "-", 3, "PROJ", "PROJ-03 项目详情"),
    ("PROJ-03-FE3", "检测进度条（SSE 驱动）：进度比例 + 一行最新动态摘要", "FE", "中", "US-5.3", "GET /api/projects/{pid}/events", "-", 4, "PROJ", "PROJ-03 项目详情"),
    ("PROJ-04-BE1", "DELETE /api/projects/{id} 接口（级联删除全部关联数据 + 磁盘文件清理）", "BE", "中", "US-2.4", "DELETE /api/projects/{id}", "Project及全部级联表", 4, "PROJ", "PROJ-04 删除项目"),
    ("PROJ-04-FE1", "删除确认弹窗 + analyzing 状态禁止删除", "FE", "小", "US-2.4", "DELETE /api/projects/{id}", "-", 2, "PROJ", "PROJ-04 删除项目"),

    # === FILE ===
    ("FILE-01-BE1", "POST /api/projects/{pid}/bidders 接口（multipart name+file?，名称唯一，触发异步解析）", "BE", "中", "US-3.1", "POST /api/projects/{pid}/bidders", "Bidder,BidDocument", 7, "FILE", "FILE-01 添加投标人"),
    ("FILE-01-FE1", "添加投标人弹窗（名称输入 + 拖拽/选择压缩包上传区域）", "FE", "中", "US-3.1", "POST /api/projects/{pid}/bidders", "-", 4, "FILE", "FILE-01 添加投标人"),
    ("FILE-01-FE2", "前端文件校验（格式 ZIP/RAR/7Z + 大小≤500MB）", "FE", "小", "US-3.1", "-", "-", 2, "FILE", "FILE-01 添加投标人"),
    ("FILE-02-BE1", "POST /api/projects/{pid}/bidders/{bid}/upload 接口（追加模式、MD5 去重、魔数校验）", "BE", "中", "US-3.2", "POST .../bidders/{bid}/upload", "BidDocument", 8, "FILE", "FILE-02 追加上传"),
    ("FILE-02-FE1", "追加上传交互（拖拽上传 + 进度条 + 去重提示）", "FE", "小", "US-3.2", "POST .../bidders/{bid}/upload", "-", 3, "FILE", "FILE-02 追加上传"),
    ("FILE-03-BE1", "GET /api/projects/{pid}/bidders/{bid}/documents 接口", "BE", "小", "US-3.3", "GET .../bidders/{bid}/documents", "BidDocument", 3, "FILE", "FILE-03 文件列表"),
    ("FILE-03-BE2", "GET /api/documents/{id}/download 接口（文件流，过期 410 Gone）", "BE", "小", "US-3.3", "GET /api/documents/{id}/download", "BidDocument", 3, "FILE", "FILE-03 文件列表"),
    ("FILE-03-FE1", "文件列表展开子表格（树形结构、角色标签、状态图标、下载按钮）", "FE", "中", "US-3.3", "上述两个 API", "-", 7, "FILE", "FILE-03 文件列表"),
    ("FILE-04-BE1", "PATCH /api/documents/{id}/role 接口", "BE", "小", "US-4.3", "PATCH /api/documents/{id}/role", "BidDocument", 2, "FILE", "FILE-04 修正文档角色"),
    ("FILE-04-FE1", "角色标签点击弹出下拉修改 + 待确认黄色高亮 + completed 提示 banner", "FE", "小", "US-4.3", "PATCH /api/documents/{id}/role", "-", 3, "FILE", "FILE-04 修正文档角色"),
    ("FILE-05-BE1", "DELETE /api/projects/{pid}/bidders/{bid} 接口（级联删除 + 项目状态回退）", "BE", "中", "US-3.4", "DELETE .../bidders/{bid}", "Bidder及级联表", 4, "FILE", "FILE-05 删除投标人"),
    ("FILE-05-FE1", "删除确认弹窗 + analyzing 状态禁止删除", "FE", "小", "US-3.4", "DELETE .../bidders/{bid}", "-", 2, "FILE", "FILE-05 删除投标人"),
    ("FILE-06-BE1", "POST /api/documents/{id}/decrypt 接口（密码解压重试）", "BE", "小", "US-4.1", "POST /api/documents/{id}/decrypt", "BidDocument", 2, "FILE", "FILE-06 加密包解密"),
    ("FILE-06-FE1", "需密码状态展示 + 密码输入弹窗", "FE", "小", "US-4.1", "POST /api/documents/{id}/decrypt", "-", 1, "FILE", "FILE-06 加密包解密"),
    ("FILE-07-BE1", "GET /api/projects/{pid}/price-rules 接口", "BE", "小", "US-4.4", "GET .../price-rules", "PriceParsingRule", 1, "FILE", "FILE-07 报价解析规则"),
    ("FILE-07-BE2", "PUT /api/projects/{pid}/price-rules/{id} 接口（修正列映射 + 触发批量重新提取）", "BE", "中", "US-4.4", "PUT .../price-rules/{id}", "PriceParsingRule,PriceItem", 3, "FILE", "FILE-07 报价解析规则"),
    ("FILE-07-FE1", "报价规则确认/修正界面（列映射下拉、sheet 选择、表头行配置）", "FE", "中", "US-4.4", "上述两个 API", "-", 4, "FILE", "FILE-07 报价解析规则"),

    # === PARSE ===
    ("PARSE-01-BE1", "递归解压引擎（ZIP/RAR/7Z，最大嵌套3层）", "BE", "大", "US-4.1", "-", "BidDocument", 4, "PARSE", "PARSE-01 解压与安全校验"),
    ("PARSE-01-BE2", "安全校验：zip bomb 防护 + Zip Slip 路径穿越防护", "BE", "中", "US-4.1", "-", "-", 3, "PARSE", "PARSE-01 解压与安全校验"),
    ("PARSE-01-BE3", "文件名编码处理（UTF-8/GBK/GB2312 自动检测）", "BE", "小", "US-4.1", "-", "-", 1, "PARSE", "PARSE-01 解压与安全校验"),
    ("PARSE-01-BE4", "文件类型识别（魔数校验）+ 分类登记", "BE", "小", "US-4.1", "-", "BidDocument", 2, "PARSE", "PARSE-01 解压与安全校验"),
    ("PARSE-02-BE1", "DOCX 文本提取（段落 + 文本框 + 表格按行合并）", "BE", "中", "US-4.2", "-", "DocumentText", 3, "PARSE", "PARSE-02 内容提取"),
    ("PARSE-02-BE2", "DOCX 页眉页脚单独提取（存入 header_footer JSON 字段）", "BE", "小", "US-4.2", "-", "DocumentText", 1, "PARSE", "PARSE-02 内容提取"),
    ("PARSE-02-BE3", "DOCX/XLSX 元数据提取（author/last_saved_by/company 等）", "BE", "中", "US-4.2", "-", "DocumentMetadata", 2, "PARSE", "PARSE-02 内容提取"),
    ("PARSE-02-BE4", "DOCX 图片提取 + MD5 + pHash 计算", "BE", "中", "US-4.2", "-", "DocumentImage", 2, "PARSE", "PARSE-02 内容提取"),
    ("PARSE-02-BE5", "XLSX 单元格数据提取 + 隐藏 sheet/列扫描（硬件信息关键词）", "BE", "中", "US-4.2", "-", "DocumentMetadata", 2, "PARSE", "PARSE-02 内容提取"),
    ("PARSE-03-BE1", "规则预分类（文件名关键词匹配 9 种角色）", "BE", "小", "US-4.3", "-", "BidDocument", 1, "PARSE", "PARSE-03 LLM角色分类"),
    ("PARSE-03-BE2", "LLM 批量分类 + 标识提取（L-1 prompt，每投标人1次调用）", "BE", "大", "US-4.3", "-", "BidDocument,Bidder", 4, "PARSE", "PARSE-03 LLM角色分类"),
    ("PARSE-03-BE3", "兜底逻辑：LLM 失败→关键词匹配角色 + 标识信息留空", "BE", "小", "US-4.3", "-", "BidDocument,Bidder", 2, "PARSE", "PARSE-03 LLM角色分类"),
    ("PARSE-04-BE1", "LLM 报价结构识别（L-2 prompt，每项目1次）", "BE", "大", "US-4.4", "-", "PriceParsingRule", 3, "PARSE", "PARSE-04 报价结构识别"),
    ("PARSE-04-BE2", "规则批量应用引擎（提取报价数据到 PriceItem）", "BE", "中", "US-4.4", "-", "PriceItem", 2, "PARSE", "PARSE-04 报价结构识别"),
    ("PARSE-04-BE3", "并发控制（asyncio.Lock）+ 自动回填已解析投标人", "BE", "中", "US-4.4", "-", "PriceParsingRule,PriceItem", 2, "PARSE", "PARSE-04 报价结构识别"),
    ("PARSE-04-BE4", "兜底：识别失败→标记待手动配置 + 批量失败→单独 LLM 重试", "BE", "小", "US-4.4", "-", "PriceParsingRule", 2, "PARSE", "PARSE-04 报价结构识别"),
    ("PARSE-05-BE1", "异步流水线编排（5步骤串行/并行）", "BE", "大", "F-PA-05", "-", "全部解析相关表", 5, "PARSE", "PARSE-05 流水线编排"),
    ("PARSE-05-BE2", "状态级联更新（BidDocument→Bidder→Project 状态转换）", "BE", "中", "F-PA-05", "-", "Project,Bidder,BidDocument", 3, "PARSE", "PARSE-05 流水线编排"),
    ("PARSE-05-BE3", "SSE 推送 parse_progress 事件", "BE", "小", "F-PA-05", "GET .../events", "-", 2, "PARSE", "PARSE-05 流水线编排"),

    # === DETECT ===
    ("DETECT-01-BE1", "POST /api/projects/{pid}/analysis/start 接口（前置校验、版本分配、AgentTask 创建）", "BE", "中", "US-5.1", "POST .../analysis/start", "AgentTask", 6, "DETECT", "DETECT-01 启动检测"),
    ("DETECT-01-FE1", "启动检测按钮逻辑（启禁条件、防重复提交、进度面板展开）", "FE", "小", "US-5.1", "POST .../analysis/start", "-", 3, "DETECT", "DETECT-01 启动检测"),
    ("DETECT-02-BE1", "Agent 调度框架（asyncio.gather 并行、各 Agent 自检、跳过/执行分支）", "BE", "大", "US-5.2", "-", "AgentTask", 4, "DETECT", "DETECT-02 Agent框架"),
    ("DETECT-02-BE2", "SSE 推送 agent_status + report_ready 事件", "BE", "小", "US-5.3", "GET .../events", "-", 2, "DETECT", "DETECT-02 Agent框架"),
    ("DETECT-03-BE1", "硬件特征码 Agent：三层提取 + 精确匹配（铁证级）", "BE", "大", "F-DA-01", "-", "PairComparison,DocumentMetadata", 4, "DETECT", "DETECT-03 硬件特征码"),
    ("DETECT-04-BE1", "错误一致性 Agent：关键词交叉初筛 + 防碰撞（短词/高频/上限100段）", "BE", "中", "F-DA-02", "-", "DocumentText,Bidder", 3, "DETECT", "DETECT-04 错误一致性"),
    ("DETECT-04-BE2", "错误一致性 Agent：LLM 深度判断（L-5，双向检查+页眉页脚，铁证级）", "BE", "大", "F-DA-02", "-", "PairComparison", 3, "DETECT", "DETECT-04 错误一致性"),
    ("DETECT-04-BE3", "错误一致性 Agent：降级模式（identity_info 空→用投标人名称，不跳过）", "BE", "小", "F-DA-02", "-", "PairComparison", 2, "DETECT", "DETECT-04 错误一致性"),
    ("DETECT-05-BE1", "文本相似度 Agent：TF-IDF + 余弦相似度（段落级，同角色对齐）", "BE", "大", "F-DA-03", "-", "PairComparison,DocumentText", 3, "DETECT", "DETECT-05 文本相似度"),
    ("DETECT-05-BE2", "文本相似度 Agent：LLM 同源性鉴别（L-4）", "BE", "中", "F-DA-03", "-", "PairComparison", 2, "DETECT", "DETECT-05 文本相似度"),
    ("DETECT-06-BE1", "价格构成 Agent：清单对齐（编码→前缀→名称→LLM→兜底跳过）", "BE", "大", "F-DA-04", "-", "PriceItem", 3, "DETECT", "DETECT-06 价格构成"),
    ("DETECT-06-BE2", "价格构成 Agent：逐项相似度 + 分布模式比对", "BE", "中", "F-DA-04", "-", "PairComparison", 2, "DETECT", "DETECT-06 价格构成"),
    ("DETECT-06-BE3", "价格构成 Agent：LLM 合理性判断（L-6）", "BE", "中", "F-DA-04", "-", "PairComparison", 2, "DETECT", "DETECT-06 价格构成"),
    ("DETECT-07-BE1", "图片复用 Agent：MD5 精确匹配 + pHash 模糊匹配", "BE", "中", "F-DA-05", "-", "PairComparison,DocumentImage", 2, "DETECT", "DETECT-07 图片复用"),
    ("DETECT-07-BE2", "图片复用 Agent：LLM 通用性判断（L-7）+ 铁证升级", "BE", "中", "F-DA-05", "-", "PairComparison", 2, "DETECT", "DETECT-07 图片复用"),
    ("DETECT-08-BE1", "语言风格 Agent：TF-IDF 预过滤 + 抽样 5-10 段", "BE", "中", "F-DA-06", "-", "DocumentText", 1, "DETECT", "DETECT-08 语言风格"),
    ("DETECT-08-BE2", "语言风格 Agent：LLM 两阶段分析（L-8，>20人自动分组）", "BE", "大", "F-DA-06", "-", "PairComparison", 3, "DETECT", "DETECT-08 语言风格"),
    ("DETECT-09-BE1", "软件特征 Agent：元数据精确匹配 + 白名单过滤（纯程序）", "BE", "小", "F-DA-07", "-", "PairComparison,DocumentMetadata", 2, "DETECT", "DETECT-09 软件特征"),
    ("DETECT-10-BE1", "操作时间 Agent：滑动窗口聚集 + 非工作时间占比", "BE", "中", "F-DA-08", "-", "OverallAnalysis,DocumentMetadata", 2, "DETECT", "DETECT-10 操作时间"),
    ("DETECT-11-BE1", "报价规律性 Agent：等差/等比/百分比数列检验", "BE", "中", "F-DA-09", "-", "OverallAnalysis,PriceItem", 2, "DETECT", "DETECT-11 报价规律"),
    ("DETECT-12-BE1", "接近限价 Agent：偏离度方差统计", "BE", "小", "F-DA-10", "-", "OverallAnalysis,PriceItem", 2, "DETECT", "DETECT-12 接近限价"),
    ("DETECT-13-BE1", "Pair 评分（加权公式）+ 铁证直判 + Project 评分", "BE", "中", "F-RP-01", "-", "PairComparison,OverallAnalysis", 3, "DETECT", "DETECT-13 综合评分"),
    ("DETECT-13-BE2", "LLM 综合研判（输入预聚合摘要→风险结论）+ 兜底", "BE", "大", "4.4.1", "-", "AnalysisReport", 3, "DETECT", "DETECT-13 综合评分"),
    ("DETECT-13-BE3", "AnalysisReport 创建 + Project 状态→completed", "BE", "小", "US-5.1", "-", "AnalysisReport,Project", 2, "DETECT", "DETECT-13 综合评分"),
    ("DETECT-14-BE1", "Agent 超时控制（5分钟/Agent, 30分钟/全局）+ Process.kill()", "BE", "中", "US-5.4", "-", "AgentTask", 3, "DETECT", "DETECT-14 超时与兜底"),
    ("DETECT-14-BE2", "LLM 调用重试（超时30秒，重试2次）+ 失败降级", "BE", "中", "US-5.4", "-", "-", 3, "DETECT", "DETECT-14 超时与兜底"),

    # === REPORT ===
    ("REPORT-01-BE1", "GET /api/projects/{pid}/reports 接口（报告列表）", "BE", "小", "US-6.1", "GET .../reports", "AnalysisReport", 1, "REPORT", "REPORT-01 报告总览"),
    ("REPORT-01-BE2", "GET /api/projects/{pid}/reports/{ver} 接口（完整报告 + 维度汇总）", "BE", "中", "US-6.1", "GET .../reports/{ver}", "AnalysisReport,PairComparison,OverallAnalysis", 3, "REPORT", "REPORT-01 报告总览"),
    ("REPORT-01-BE3", "GET /api/projects/{pid}/reports/{ver}/matrix 接口", "BE", "小", "US-6.1", "GET .../reports/{ver}/matrix", "PairComparison", 2, "REPORT", "REPORT-01 报告总览"),
    ("REPORT-01-FE1", "报告页 Tab1 概要：风险徽章+LLM结论+雷达图+热力图+汇总表", "FE", "大", "US-6.1", "上述3个 API", "-", 8, "REPORT", "REPORT-01 报告总览"),
    ("REPORT-02-BE1", "GET .../reports/{ver}/dimensions/{dim} 接口（区分 pair/overall 响应）", "BE", "中", "US-6.2", "GET .../dimensions/{dim}", "PairComparison,OverallAnalysis", 3, "REPORT", "REPORT-02 维度明细"),
    ("REPORT-02-FE1", "Tab2 维度明细：Master-Detail 布局", "FE", "中", "US-6.2", "上述 API", "-", 7, "REPORT", "REPORT-02 维度明细"),
    ("REPORT-03-BE1", "GET .../reports/{ver}/pairs 接口（?bidder_a=&bidder_b=）", "BE", "小", "US-6.3", "GET .../pairs", "PairComparison", 2, "REPORT", "REPORT-03 投标人对详情"),
    ("REPORT-03-FE1", "Tab3 投标人对：下拉选择+维度得分表+对比/查看链接", "FE", "中", "US-6.3", "上述 API", "-", 5, "REPORT", "REPORT-03 投标人对详情"),
    ("REPORT-04-BE1", "GET .../reports/{ver}/logs 接口", "BE", "小", "US-6.4", "GET .../logs", "AgentTask", 1, "REPORT", "REPORT-04 检测日志"),
    ("REPORT-04-FE1", "Tab4 检测日志：时间线表格+状态筛选+展开详情", "FE", "中", "US-6.4", "上述 API", "-", 5, "REPORT", "REPORT-04 检测日志"),
    ("REPORT-05-BE1", "POST .../reports/{ver}/review 接口", "BE", "小", "US-6.5", "POST .../review", "AnalysisReport", 4, "REPORT", "REPORT-05 人工复核"),
    ("REPORT-05-FE1", "复核弹窗（风险等级下拉+意见文本域）+ 复核状态展示", "FE", "小", "US-6.5", "上述 API", "-", 3, "REPORT", "REPORT-05 人工复核"),
    ("REPORT-06-BE1", "GET .../reports/{ver}/export 接口（python-docx 生成 DOCX）", "BE", "大", "US-6.6", "GET .../export", "AnalysisReport,PairComparison,OverallAnalysis", 5, "REPORT", "REPORT-06 Word导出"),
    ("REPORT-06-FE1", "导出 Word 按钮 + 浏览器下载触发", "FE", "小", "US-6.6", "上述 API", "-", 2, "REPORT", "REPORT-06 Word导出"),

    # === COMPARE ===
    ("COMPARE-01-BE1", "GET .../compare/text 接口（段落+相似度匹配，支持 version）", "BE", "中", "US-7.1", "GET .../compare/text", "DocumentText,PairComparison", 2, "COMPARE", "COMPARE-01 文本对比"),
    ("COMPARE-01-FE1", "文本对比页面：左右分栏、同步滚动、相似段落高亮、角色切换", "FE", "大", "US-7.1", "上述 API", "-", 6, "COMPARE", "COMPARE-01 文本对比"),
    ("COMPARE-02-BE1", "GET .../compare/price 接口（对齐后报价项+偏差，支持 version）", "BE", "中", "US-7.2", "GET .../compare/price", "PriceItem,PairComparison", 2, "COMPARE", "COMPARE-02 报价对比"),
    ("COMPARE-02-FE1", "报价对比页面：逐项对比表格、偏差色标、排序", "FE", "中", "US-7.2", "上述 API", "-", 5, "COMPARE", "COMPARE-02 报价对比"),
    ("COMPARE-03-BE1", "GET .../compare/metadata 接口（元数据矩阵+白名单，支持 version）", "BE", "小", "US-7.3", "GET .../compare/metadata", "DocumentMetadata", 2, "COMPARE", "COMPARE-03 元数据对比"),
    ("COMPARE-03-FE1", "元数据对比页面：矩阵表格、匹配高亮、白名单标灰", "FE", "中", "US-7.3", "上述 API", "-", 5, "COMPARE", "COMPARE-03 元数据对比"),

    # === ADMIN ===
    ("ADMIN-01-BE1", "GET /api/admin/users 接口", "BE", "小", "US-8.1", "GET /api/admin/users", "User", 2, "ADMIN", "ADMIN-01 用户列表"),
    ("ADMIN-01-FE1", "用户管理页面（/admin/users）：表格+搜索", "FE", "小", "US-8.1", "GET /api/admin/users", "-", 2, "ADMIN", "ADMIN-01 用户列表"),
    ("ADMIN-02-BE1", "POST /api/admin/users 接口（用户名唯一、密码规则）", "BE", "小", "US-8.2", "POST /api/admin/users", "User", 4, "ADMIN", "ADMIN-02 创建用户"),
    ("ADMIN-02-FE1", "创建用户表单弹窗", "FE", "小", "US-8.2", "POST /api/admin/users", "-", 2, "ADMIN", "ADMIN-02 创建用户"),
    ("ADMIN-03-BE1", "PATCH /api/admin/users/{id} 接口（禁用/启用/角色，不可禁用自己）", "BE", "小", "US-8.3", "PATCH /api/admin/users/{id}", "User", 3, "ADMIN", "ADMIN-03 禁用启用"),
    ("ADMIN-03-FE1", "用户行操作按钮（禁用/启用切换、角色修改）", "FE", "小", "US-8.3", "PATCH /api/admin/users/{id}", "-", 1, "ADMIN", "ADMIN-03 禁用启用"),
    ("ADMIN-04-BE1", "GET/PUT /api/admin/rules 接口（JSON 配置读写+校验）", "BE", "中", "US-9.1", "GET/PUT /api/admin/rules", "SystemConfig", 5, "ADMIN", "ADMIN-04 规则配置"),
    ("ADMIN-04-BE2", "系统启动时加载默认配置到 SystemConfig 表", "BE", "小", "US-9.1", "-", "SystemConfig", 1, "ADMIN", "ADMIN-04 规则配置"),
    ("ADMIN-04-FE1", "规则配置页面（/admin/rules）：JSON 配置表单+恢复默认", "FE", "中", "US-9.1", "GET/PUT /api/admin/rules", "-", 3, "ADMIN", "ADMIN-04 规则配置"),

    # === INFRA ===
    ("INFRA-01-BE1", "SQLAlchemy 模型定义（14 张表）", "BE", "大", "TS-4", "-", "全部14张表", 5, "INFRA", "INFRA-01 数据库模型"),
    ("INFRA-01-BE2", "Alembic 初始迁移脚本", "BE", "小", "TS-4", "-", "-", 3, "INFRA", "INFRA-01 数据库模型"),
    ("INFRA-02-BE1", "asyncio + ProcessPoolExecutor 封装", "BE", "中", "TS-1", "-", "-", 2, "INFRA", "INFRA-02 异步任务"),
    ("INFRA-02-BE2", "超时机制（asyncio.wait_for + Process.kill()）", "BE", "中", "TS-1", "-", "-", 2, "INFRA", "INFRA-02 异步任务"),
    ("INFRA-02-BE3", "进程重启恢复逻辑", "BE", "中", "TS-1", "-", "Project,Bidder,BidDocument,AgentTask", 4, "INFRA", "INFRA-02 异步任务"),
    ("INFRA-03-BE1", "SSE 端点（EventSourceResponse，3种事件类型）", "BE", "中", "TS-2", "GET .../events", "-", 4, "INFRA", "INFRA-03 SSE推送"),
    ("INFRA-03-BE2", "SSE 重连恢复接口 GET .../analysis/status", "BE", "小", "TS-2", "GET .../analysis/status", "AgentTask", 2, "INFRA", "INFRA-03 SSE推送"),
    ("INFRA-03-FE1", "EventSource 客户端封装（连接/断线重连/事件分发/状态恢复）", "FE", "中", "TS-2", "上述2个 API", "-", 3, "INFRA", "INFRA-03 SSE推送"),
    ("INFRA-04-BE1", "LLM 统一调用接口（模型无关，OpenAI 兼容 API）", "BE", "中", "TS-3", "-", "-", 2, "INFRA", "INFRA-04 LLM适配层"),
    ("INFRA-04-BE2", "重试机制（超时30秒，重试2次）+ JSON 容错解析", "BE", "中", "TS-3", "-", "-", 3, "INFRA", "INFRA-04 LLM适配层"),
    ("INFRA-04-BE3", "API Key 环境变量加载 + Prompt 模板管理（Python/YAML）", "BE", "小", "TS-3", "-", "-", 2, "INFRA", "INFRA-04 LLM适配层"),
    ("INFRA-05-FE1", "路由配置（react-router-dom，11个页面）", "FE", "小", "TS-5", "-", "-", 1, "INFRA", "INFRA-05 前端基础设施"),
    ("INFRA-05-FE2", "UI 框架集成（Ant Design 主题+全局 Layout）", "FE", "中", "TS-5", "-", "-", 3, "INFRA", "INFRA-05 前端基础设施"),
    ("INFRA-05-FE3", "状态管理（zustand）+ API 客户端封装（axios）", "FE", "中", "TS-5", "-", "-", 3, "INFRA", "INFRA-05 前端基础设施"),
    ("INFRA-05-FE4", "图表库集成（ECharts：雷达图+热力图）", "FE", "小", "TS-5", "-", "-", 1, "INFRA", "INFRA-05 前端基础设施"),
    ("INFRA-06-BE1", "定时清理任务（每日执行，清理超期原始文件）", "BE", "中", "TS-6", "-", "BidDocument", 4, "INFRA", "INFRA-06 数据生命周期"),
    ("INFRA-06-BE2", "清理日志记录 + 过期下载返回已清理提示", "BE", "小", "TS-6", "-", "-", 2, "INFRA", "INFRA-06 数据生命周期"),
]

# ============================================================
# 样式定义
# ============================================================
HEADER_FONT = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
MODULE_FONT = Font(name="微软雅黑", bold=True, size=11)
MODULE_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
FEATURE_FONT = Font(name="微软雅黑", bold=True, size=10)
FEATURE_FILL = PatternFill(start_color="E9EFF7", end_color="E9EFF7", fill_type="solid")
NORMAL_FONT = Font(name="微软雅黑", size=10)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

SIZE_FILL = {
    "小": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "中": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "大": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
}

# ============================================================
# Sheet 1: 详细清单
# ============================================================
ws = wb.active
ws.title = "详细清单"

HEADERS = ["编号", "模块", "功能项", "任务描述", "归属", "工作量", "关联US", "关联API", "关联表", "AC数"]
COL_WIDTHS = [16, 14, 22, 60, 6, 8, 10, 35, 30, 6]

for col_idx, (header, width) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = THIN_BORDER
    ws.column_dimensions[get_column_letter(col_idx)].width = width

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"

row = 2
prev_module = ""
prev_feature = ""

for t in TASKS:
    tid, desc, scope, size, us, api, tables, ac, module, feature = t

    # 模块分组行
    module_label = f"{module} - {dict(MODULES)[module]}" if module in dict(MODULES) else module
    if module != prev_module:
        for col_idx in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=row, column=col_idx)
            cell.fill = MODULE_FILL
            cell.font = MODULE_FONT
            cell.border = THIN_BORDER
        ws.cell(row=row, column=1, value=module_label)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(HEADERS))
        row += 1
        prev_module = module
        prev_feature = ""

    # 功能项分组行
    if feature != prev_feature:
        for col_idx in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=row, column=col_idx)
            cell.fill = FEATURE_FILL
            cell.font = FEATURE_FONT
            cell.border = THIN_BORDER
        ws.cell(row=row, column=3, value=feature)
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=len(HEADERS))
        row += 1
        prev_feature = feature

    # 任务行
    values = [tid, "", "", desc, scope, size, us, api, tables, ac]
    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.font = NORMAL_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    # 工作量色标
    size_cell = ws.cell(row=row, column=6)
    if size in SIZE_FILL:
        size_cell.fill = SIZE_FILL[size]
    size_cell.alignment = Alignment(horizontal="center", vertical="center")
    # 归属居中
    ws.cell(row=row, column=5).alignment = Alignment(horizontal="center", vertical="center")
    # AC数居中
    ws.cell(row=row, column=10).alignment = Alignment(horizontal="center", vertical="center")
    row += 1

# ============================================================
# Sheet 2: 模块汇总
# ============================================================
ws2 = wb.create_sheet("模块汇总")

SUMMARY_HEADERS = ["模块", "功能项数", "任务总数", "BE任务", "FE任务", "小(S)", "中(M)", "大(L)", "预估人日(低)", "预估人日(高)"]
for col_idx, header in enumerate(SUMMARY_HEADERS, 1):
    cell = ws2.cell(row=1, column=col_idx, value=header)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = THIN_BORDER

ws2.column_dimensions["A"].width = 20
for c in "BCDEFGHIJ":
    ws2.column_dimensions[c].width = 12

summary_row = 2
total_features = total_tasks = total_be = total_fe = total_s = total_m = total_l = total_low = total_high = 0

for mod_code, mod_name in MODULES:
    mod_tasks = [t for t in TASKS if t[8] == mod_code]
    features = len(set(t[9] for t in mod_tasks))
    count = len(mod_tasks)
    be = sum(1 for t in mod_tasks if t[2] == "BE")
    fe = sum(1 for t in mod_tasks if t[2] == "FE")
    s = sum(1 for t in mod_tasks if t[3] == "小")
    m = sum(1 for t in mod_tasks if t[3] == "中")
    l = sum(1 for t in mod_tasks if t[3] == "大")
    low = s * 1 + m * 3 + l * 5
    high = s * 2 + m * 5 + l * 10

    total_features += features
    total_tasks += count
    total_be += be
    total_fe += fe
    total_s += s
    total_m += m
    total_l += l
    total_low += low
    total_high += high

    values = [f"{mod_code} {mod_name}", features, count, be, fe, s, m, l, low, high]
    for col_idx, val in enumerate(values, 1):
        cell = ws2.cell(row=summary_row, column=col_idx, value=val)
        cell.font = NORMAL_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")
    summary_row += 1

# 合计行
totals = ["合计", total_features, total_tasks, total_be, total_fe, total_s, total_m, total_l, total_low, total_high]
for col_idx, val in enumerate(totals, 1):
    cell = ws2.cell(row=summary_row, column=col_idx, value=val)
    cell.font = Font(name="微软雅黑", bold=True, size=11)
    cell.fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal="center", vertical="center")

# ============================================================
# 保存
# ============================================================
output_path = r"D:\documentcheck\docs\task-checklist.xlsx"
wb.save(output_path)
print(f"Excel 生成成功: {output_path}")
print(f"任务总数: {len(TASKS)}")
print(f"模块数: {len(MODULES)}")
print(f"功能项数: {total_features}")
print(f"预估人日: {total_low}-{total_high}")
